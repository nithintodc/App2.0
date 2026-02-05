"""Export functions for Excel and date exports"""
import pandas as pd
import streamlit as st
from datetime import datetime
from pathlib import Path
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from config import ROOT_DIR
from gdrive_utils import get_drive_manager
from utils import normalize_store_id_column, filter_master_file_by_date_range, UE_DATE_COLUMN_VARIATIONS, DD_DATE_COLUMN_VARIATIONS
from table_generation import create_summary_tables
from data_processing import get_last_year_dates


def export_to_excel(dd_table1, dd_table2, ue_table1, ue_table2, 
                     dd_sales_df, dd_payouts_df, dd_orders_df, dd_new_customers_df,
                     ue_sales_df, ue_payouts_df, ue_orders_df, ue_new_customers_df,
                     dd_selected_stores, ue_selected_stores,
                     combined_summary1, combined_summary2, combined_store_table1, combined_store_table2,
                     corporate_todc_table=None, promotion_table=None, sponsored_table=None,
                     summary_metrics_table=None, store_ids_markups_table=None, operator_name=None,
                     sales_pre_post_table=None, sales_yoy_table=None, payouts_pre_post_table=None, payouts_yoy_table=None):
    """Export all tables to an Excel file with sheets: Summary Tables, Store-Level Tables, and Corporate vs TODC"""
    # Use temp directory for file creation (will be downloaded, not saved to disk)
    import tempfile
    temp_dir = Path(tempfile.gettempdir())
    outputs_dir = temp_dir / "streamlit_exports"
    outputs_dir.mkdir(exist_ok=True)
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Get summary tables
    dd_summary1, dd_summary2 = None, None
    ue_summary1, ue_summary2 = None, None
    if dd_sales_df is not None and dd_payouts_df is not None and dd_orders_df is not None:
        dd_summary1, dd_summary2 = create_summary_tables(dd_sales_df, dd_payouts_df, dd_orders_df, dd_new_customers_df, dd_selected_stores, is_ue=False)
    if ue_sales_df is not None and ue_payouts_df is not None and ue_orders_df is not None:
        ue_summary1, ue_summary2 = create_summary_tables(ue_sales_df, ue_payouts_df, ue_orders_df, ue_new_customers_df, ue_selected_stores, is_ue=True)
    
    # Sheet 1: Summary Tables
    ws_summary = wb.create_sheet("Summary Tables")
    current_row = 1
    
    def add_table_to_sheet(ws, table_name, df, start_row, start_col=1):
        """Add a table with name header to the sheet and format it. start_col is 1-based column for table placement."""
        if df is None or df.empty:
            return start_row
        # Add table name
        ws.cell(row=start_row, column=start_col, value=table_name)
        ws.cell(row=start_row, column=start_col).font = Font(bold=True, size=12)
        start_row += 1
        # Add table data
        # Only reset index when it has a known name we want as a column (don't export default RangeIndex for slot/markup tables)
        if df.index.name in ['Store ID', 'Metric', 'Campaign', 'Is Self Serve Campaign']:
            df_display = df.reset_index()
        else:
            df_display = df.copy()
        
        # Write header row
        for col_idx, col_name in enumerate(df_display.columns, start=1):
            cell = ws.cell(row=start_row, column=start_col + col_idx - 1, value=col_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        start_row += 1
        
        # Write data rows
        for row_idx, row_data in df_display.iterrows():
            # Check if this row is for Orders or New Customers (for summary tables)
            is_orders_row = False
            is_new_customers_row = False
            metric_value = ''  # Initialize metric_value
            
            # Try to get metric value from column first (after reset_index, Metric becomes a column)
            if 'Metric' in df_display.columns:
                try:
                    metric_val = row_data['Metric']
                    metric_value = str(metric_val) if pd.notna(metric_val) else ''
                except (KeyError, IndexError, TypeError):
                    metric_value = ''
                
                if metric_value == 'Orders':
                    is_orders_row = True
                elif metric_value == 'New Customers':
                    is_new_customers_row = True
            # If Metric is still in the index (wasn't reset), use row_idx directly
            elif df_display.index.name == 'Metric':
                metric_value = str(row_idx) if pd.notna(row_idx) else ''
            # Last resort: if row_idx itself is the metric name (string)
            elif isinstance(row_idx, str):
                metric_value = str(row_idx)
            
            for col_idx, col_name in enumerate(df_display.columns, start=1):
                value = row_data[col_name]
                cell = ws.cell(row=start_row, column=start_col + col_idx - 1, value=value)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Format based on column name and row type
                if 'Growth%' in col_name or 'YoY%' in col_name:
                    # Format as percentage with % symbol
                    if isinstance(value, (int, float)):
                        cell.value = f"{value:.1f}%"
                elif col_name in ['Store ID', 'Metric', 'Campaign', 'Is Self Serve Campaign', 'Merchant Store IDs', 'Markups']:
                    # Keep as is (text)
                    pass
                elif col_name == 'Orders':
                    # Orders: format as integer with comma separators
                    if isinstance(value, (int, float)):
                        cell.value = f"{int(round(value)):,}"
                elif col_name in ['Sales', 'Spend', 'Cost per Order']:
                    # Sales, Spend, Cost per Order: format as dollar amount
                    if isinstance(value, (int, float)):
                        cell.value = f"${value:,.2f}"
                elif col_name == 'ROAS':
                    # ROAS: format as decimal
                    if isinstance(value, (int, float)):
                        cell.value = f"{value:.2f}"
                elif is_orders_row or is_new_customers_row:
                    # Orders and New Customers rows: format as integer with comma separators (no decimals, no dollar sign)
                    if isinstance(value, (int, float)):
                        cell.value = f"{int(round(value)):,}"
                elif metric_value == 'Profitability':
                    # Profitability: format as percentage
                    if isinstance(value, (int, float)):
                        cell.value = f"{value:.1f}%"
                elif metric_value == 'AOV':
                    # AOV: format as dollar amount
                    if isinstance(value, (int, float)):
                        cell.value = f"${value:,.1f}"
                else:
                    # Format as dollar amount
                    if isinstance(value, (int, float)):
                        cell.value = f"${value:,.1f}"
            
            start_row += 1
        
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(df_display.columns, start=1):
            c = start_col + col_idx - 1
            col_series = df_display[col_name].astype(str)
            max_len = int(col_series.str.len().max()) if len(col_series) else 0
            max_length = max(len(str(col_name)), max_len)
            ws.column_dimensions[get_column_letter(c)].width = min(max_length + 2, 50)
        
        return start_row + 1  # Add blank row after table
    
    # Add Summary Metrics table first
    if summary_metrics_table is not None and not summary_metrics_table.empty:
        current_row = add_table_to_sheet(ws_summary, "Summary Metrics", summary_metrics_table, current_row)
    
    # Add Merchant Store IDs / Markups table beside Summary Metrics with 5 columns gap (export only)
    if store_ids_markups_table is not None and not store_ids_markups_table.empty:
        # Summary Metrics has 2 columns; gap = 5 columns; so start at column 1+2+5 = 8
        add_table_to_sheet(ws_summary, "Merchant Store IDs / Markups", store_ids_markups_table, start_row=1, start_col=8)
    
    # Add Combined Table 1
    if combined_summary1 is not None:
        current_row = add_table_to_sheet(ws_summary, "Combined Table 1: Current Year Pre vs Post Analysis", combined_summary1, current_row)
    
    # Add Combined Table 2 (YoY) - second position
    if combined_summary2 is not None:
        current_row = add_table_to_sheet(ws_summary, "Combined Table 2: Year-over-Year Analysis", combined_summary2, current_row)
    
    # Add DD Table 1
    if dd_summary1 is not None:
        current_row = add_table_to_sheet(ws_summary, "DoorDash Table 1: Current Year Pre vs Post Analysis", dd_summary1, current_row)
    
    # Add DD Table 2
    if dd_summary2 is not None:
        current_row = add_table_to_sheet(ws_summary, "DoorDash Table 2: Year-over-Year Analysis", dd_summary2, current_row)
    
    # Add UE Table 1
    if ue_summary1 is not None:
        current_row = add_table_to_sheet(ws_summary, "UberEats Table 1: Current Year Pre vs Post Analysis", ue_summary1, current_row)
    
    # Add UE Table 2
    if ue_summary2 is not None:
        current_row = add_table_to_sheet(ws_summary, "UberEats Table 2: Year-over-Year Analysis", ue_summary2, current_row)
    
    # Sheet 2: Store-Level Tables
    ws_store = wb.create_sheet("Store-Level Tables")
    current_row = 1
    
    # Add Combined Store Table 1
    if combined_store_table1 is not None:
        current_row = add_table_to_sheet(ws_store, "Combined Table 1: Current Year Pre vs Post Analysis (Store-Level)", combined_store_table1, current_row)
    
    # Add Combined Store Table 2 (YoY) - second position
    if combined_store_table2 is not None:
        current_row = add_table_to_sheet(ws_store, "Combined Table 2: Year-over-Year Analysis (Store-Level)", combined_store_table2, current_row)
    
    # Add DD Store Table 1
    if dd_table1 is not None:
        current_row = add_table_to_sheet(ws_store, "DoorDash Table 1: Current Year Pre vs Post Analysis (Store-Level)", dd_table1, current_row)
    
    # Add DD Store Table 2
    if dd_table2 is not None:
        current_row = add_table_to_sheet(ws_store, "DoorDash Table 2: Year-over-Year Analysis (Store-Level)", dd_table2, current_row)
    
    # Add UE Store Table 1
    if ue_table1 is not None:
        current_row = add_table_to_sheet(ws_store, "UberEats Table 1: Current Year Pre vs Post Analysis (Store-Level)", ue_table1, current_row)
    
    # Add UE Store Table 2
    if ue_table2 is not None:
        current_row = add_table_to_sheet(ws_store, "UberEats Table 2: Year-over-Year Analysis (Store-Level)", ue_table2, current_row)
    
    # Sheet 3: Corporate vs TODC Tables
    if corporate_todc_table is not None and not corporate_todc_table.empty:
        ws_corporate = wb.create_sheet("Corporate vs TODC")
        current_row = 1
        
        # Add Combined Corporate vs TODC table
        # Prepare the table for export (reset index to include Campaign as column)
        corporate_export = corporate_todc_table.copy()
        corporate_export.index.name = 'Campaign'
        corporate_export = corporate_export.reset_index()
        corporate_export['Campaign'] = corporate_export['Campaign'].apply(
            lambda x: 'Corporate' if x == False else ('TODC' if x == True else str(x))
        )
        corporate_export = corporate_export.set_index('Campaign')
        
        current_row = add_table_to_sheet(ws_corporate, "Combined: Corporate vs TODC", corporate_export, current_row)
        
        # Add Promotion table if available
        if promotion_table is not None and not promotion_table.empty:
            promo_export = promotion_table.copy()
            promo_export.index.name = 'Campaign'
            promo_export = promo_export.reset_index()
            promo_export['Campaign'] = promo_export['Campaign'].apply(
                lambda x: 'Corporate' if x == False else ('TODC' if x == True else str(x))
            )
            promo_export = promo_export.set_index('Campaign')
            current_row = add_table_to_sheet(ws_corporate, "Promotion: Corporate vs TODC", promo_export, current_row)
        
        # Add Sponsored Listing table if available
        if sponsored_table is not None and not sponsored_table.empty:
            sponsored_export = sponsored_table.copy()
            sponsored_export.index.name = 'Campaign'
            sponsored_export = sponsored_export.reset_index()
            sponsored_export['Campaign'] = sponsored_export['Campaign'].apply(
                lambda x: 'Corporate' if x == False else ('TODC' if x == True else str(x))
            )
            sponsored_export = sponsored_export.set_index('Campaign')
            current_row = add_table_to_sheet(ws_corporate, "Sponsored Listing: Corporate vs TODC", sponsored_export, current_row)
    
    # Add Slot-based Analysis sheet
    if sales_pre_post_table is not None or sales_yoy_table is not None or payouts_pre_post_table is not None or payouts_yoy_table is not None:
        ws_slots = wb.create_sheet("Slot-based Analysis")
        current_row = 1
        
        # Add Table 1: Sales Pre/Post
        if sales_pre_post_table is not None and not sales_pre_post_table.empty:
            current_row = add_table_to_sheet(ws_slots, "Table 1: Sales - Pre vs Post", sales_pre_post_table, current_row)
            current_row += 2
        
        # Add Table 2: Sales YoY
        if sales_yoy_table is not None and not sales_yoy_table.empty:
            current_row = add_table_to_sheet(ws_slots, "Table 2: Sales - Year over Year", sales_yoy_table, current_row)
            current_row += 2
        
        # Add Table 3: Payouts Pre/Post
        if payouts_pre_post_table is not None and not payouts_pre_post_table.empty:
            current_row = add_table_to_sheet(ws_slots, "Table 3: Payouts - Pre vs Post", payouts_pre_post_table, current_row)
            current_row += 2
        
        # Add Table 4: Payouts YoY
        if payouts_yoy_table is not None and not payouts_yoy_table.empty:
            current_row = add_table_to_sheet(ws_slots, "Table 4: Payouts - Year over Year", payouts_yoy_table, current_row)
    
    # Generate filename with timestamp (use operator name if provided)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    tag = (operator_name.strip() if operator_name and isinstance(operator_name, str) and operator_name.strip() else None)
    filename = f"{tag}_analysis_export_{timestamp}.xlsx" if tag else f"analysis_export_{timestamp}.xlsx"
    filepath = outputs_dir / filename
    
    # Save workbook
    wb.save(filepath)
    
    # Read file as bytes for download
    with open(filepath, 'rb') as f:
        file_bytes = f.read()
    
    # Upload to Google Drive in "cloud-app-uploads" folder
    try:
        drive_manager = get_drive_manager()
        if drive_manager:
            # Upload to "cloud-app-uploads" folder
            upload_result = drive_manager.upload_file_to_subfolder(
                file_path=filepath,
                root_folder_name="cloud-app-uploads",
                subfolder_name="outputs",
                file_name=filename
            )
            st.success(f"✅ **Export successful!** Excel file ready for download and uploaded to Google Drive.")
            st.info(f"📤 File uploaded to Google Drive: [{upload_result['file_name']}]({upload_result['webViewLink']})")
    except Exception as e:
        st.warning(f"⚠️ Google Drive upload failed: {str(e)}")
    
    # Return file bytes and filename for download
    return file_bytes, filename


def create_date_export(dd_pre_24_path, dd_post_24_path, dd_pre_25_path, dd_post_25_path,
                      ue_pre_24_path, ue_post_24_path, ue_pre_25_path, ue_post_25_path,
                      dd_selected_stores, ue_selected_stores):
    """Create date pivot tables with Store IDs as columns and Sales, Payouts, Orders as values
    Processes only dd-pre/post and ue-pre/post files (8 files total)
    Returns a dictionary with file names as keys and dictionaries of Sales, Payouts, Orders as values
    Each file gets 3 separate pivot tables (Sales, Payouts, Orders)
    """
    
    def process_dd_file_for_date_export(file_path, selected_stores):
        """Process DD file and return data pivoted by date"""
        try:
            df = pd.read_csv(file_path)
            df.columns = df.columns.str.strip()
            
            # Use "Timestamp local date" for DD
            date_col = 'Timestamp local date'
            store_col = 'Merchant store ID'
            sales_col = 'Subtotal'
            
            # Determine payout column
            if '24' in file_path.name:
                payout_col = 'Net total (for historical reference only)'
            else:
                payout_col = 'Net total'
            
            order_col = 'DoorDash order ID'
            
            if date_col not in df.columns or store_col not in df.columns:
                st.warning(f"Missing required columns in {file_path.name}. Looking for '{date_col}' and '{store_col}'. Found: {list(df.columns)[:10]}")
                return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            
            # Check for payout column - try both names if needed
            if payout_col not in df.columns:
                # Try the other column name as fallback
                if payout_col == 'Net total':
                    payout_col = 'Net total (for historical reference only)'
                else:
                    payout_col = 'Net total'
                
                if payout_col not in df.columns:
                    st.warning(f"Payout column not found in {file_path.name}. Tried 'Net total' and 'Net total (for historical reference only)'. Available columns: {list(df.columns)[:10]}")
                    return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            
            # Process ALL data - no filtering by selected stores for date export
            # Convert date
            df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
            df = df.dropna(subset=[date_col, store_col])
            
            if len(df) == 0:
                st.warning(f"No valid data after date conversion in {file_path.name}")
                return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            
            # Convert to numeric
            df[sales_col] = pd.to_numeric(df[sales_col], errors='coerce')
            df[payout_col] = pd.to_numeric(df[payout_col], errors='coerce')
            
            # Aggregate by Date and Store ID
            if len(df) == 0:
                return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            
            sales_pivot = df.groupby([date_col, store_col])[sales_col].sum().reset_index()
            payouts_pivot = df.groupby([date_col, store_col])[payout_col].sum().reset_index()
            orders_pivot = df.groupby([date_col, store_col])[order_col].nunique().reset_index()
            
            # Check if we have data to pivot
            if len(sales_pivot) == 0:
                return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            
            # Convert Store ID to string for consistent column names
            sales_pivot[store_col] = sales_pivot[store_col].astype(str)
            payouts_pivot[store_col] = payouts_pivot[store_col].astype(str)
            orders_pivot[store_col] = orders_pivot[store_col].astype(str)
            
            # Pivot: Date as index, Store ID as columns
            sales_pivot_table = sales_pivot.pivot_table(index=date_col, columns=store_col, values=sales_col, aggfunc='sum', fill_value=0)
            payouts_pivot_table = payouts_pivot.pivot_table(index=date_col, columns=store_col, values=payout_col, aggfunc='sum', fill_value=0)
            orders_pivot_table = orders_pivot.pivot_table(index=date_col, columns=store_col, values=order_col, aggfunc='sum', fill_value=0)
            
            return sales_pivot_table, payouts_pivot_table, orders_pivot_table
        except Exception as e:
            st.error(f"Error processing {file_path.name} for date export: {str(e)}")
            import traceback
            with st.expander(f"Error details for {file_path.name}"):
                st.code(traceback.format_exc())
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    def process_ue_file_for_date_export(file_path, selected_stores):
        """Process UE file and return data pivoted by date"""
        try:
            df = pd.read_csv(file_path, skiprows=[0], header=0)
            df.columns = df.columns.str.strip()
            
            # Normalize store ID column (check for both 'Store ID' and 'Shop ID')
            df, store_col = normalize_store_id_column(df)
            
            # For UE files: hardcode to 9th column (index 8) as Order Date
            if len(df.columns) > 8:
                date_col = df.columns[8]
            else:
                st.warning(f"UE file {file_path.name} has fewer than 9 columns. Available columns: {list(df.columns)}")
                return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            
            sales_col = 'Sales (excl. tax)'
            payout_col = 'Total payout'
            order_col = 'Order ID'
            
            if date_col not in df.columns or store_col is None or store_col not in df.columns:
                st.warning(f"Missing required columns in {file_path.name}. Looking for '{date_col}' and 'Store ID' or 'Shop ID'. Found: {list(df.columns)[:10]}")
                return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            
            # Process ALL data - no filtering by selected stores for date export
            # Convert date - Store original values before parsing
            original_dates = df[date_col].copy()
            # UE files always use MM/DD/YYYY format
            df[date_col] = pd.to_datetime(df[date_col], format='%m/%d/%Y', errors='coerce')
            # Fall back to auto parsing only if format parsing fails
            if df[date_col].isna().any():
                mask_na = df[date_col].isna()
                df.loc[mask_na, date_col] = pd.to_datetime(original_dates.loc[mask_na], errors='coerce')
            df = df.dropna(subset=[date_col, store_col])
            
            if len(df) == 0:
                st.warning(f"No valid data after date conversion in {file_path.name}")
                return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
            
            # Convert to numeric
            df[sales_col] = pd.to_numeric(df[sales_col], errors='coerce')
            df[payout_col] = pd.to_numeric(df[payout_col], errors='coerce')
            
            # Aggregate by Date and Store ID
            sales_pivot = df.groupby([date_col, store_col])[sales_col].sum().reset_index()
            payouts_pivot = df.groupby([date_col, store_col])[payout_col].sum().reset_index()
            orders_pivot = df.groupby([date_col, store_col])[order_col].nunique().reset_index()
            
            # Convert Store ID to string for consistent column names
            sales_pivot[store_col] = sales_pivot[store_col].astype(str)
            payouts_pivot[store_col] = payouts_pivot[store_col].astype(str)
            orders_pivot[store_col] = orders_pivot[store_col].astype(str)
            
            # Pivot: Date as index, Store ID as columns
            sales_pivot_table = sales_pivot.pivot_table(index=date_col, columns=store_col, values=sales_col, aggfunc='sum', fill_value=0)
            payouts_pivot_table = payouts_pivot.pivot_table(index=date_col, columns=store_col, values=payout_col, aggfunc='sum', fill_value=0)
            orders_pivot_table = orders_pivot.pivot_table(index=date_col, columns=store_col, values=order_col, aggfunc='sum', fill_value=0)
            
            return sales_pivot_table, payouts_pivot_table, orders_pivot_table
        except Exception as e:
            st.error(f"Error processing {file_path.name} for date export: {str(e)}")
            import traceback
            with st.expander(f"Error details for {file_path.name}"):
                st.code(traceback.format_exc())
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    # Process all 8 files separately - each file gets its own entry
    result = {}
    
    # Process DD files
    dd_files = [
        (dd_pre_24_path, 'DD_PRE_24', dd_selected_stores),
        (dd_post_24_path, 'DD_POST_24', dd_selected_stores),
        (dd_pre_25_path, 'DD_PRE_25', dd_selected_stores),
        (dd_post_25_path, 'DD_POST_25', dd_selected_stores),
    ]
    
    for file_path, file_key, selected_stores in dd_files:
        if not file_path.exists():
            st.warning(f"File not found: {file_path}")
            continue
        try:
            sales, payouts, orders = process_dd_file_for_date_export(file_path, selected_stores)
            # Always add the file entry if at least one metric has data
            if not sales.empty or not payouts.empty or not orders.empty:
                result[file_key] = {
                    'Sales': sales if not sales.empty else pd.DataFrame(),
                    'Payouts': payouts if not payouts.empty else pd.DataFrame(),
                    'Orders': orders if not orders.empty else pd.DataFrame()
                }
            else:
                st.warning(f"No data extracted from {file_path.name} - all pivot tables are empty")
        except Exception as e:
            st.error(f"Error processing {file_path.name}: {str(e)}")
            import traceback
            with st.expander(f"Error details for {file_path.name}"):
                st.code(traceback.format_exc())
    
    # Process UE files
    ue_files = [
        (ue_pre_24_path, 'UE_PRE_24', ue_selected_stores),
        (ue_post_24_path, 'UE_POST_24', ue_selected_stores),
        (ue_pre_25_path, 'UE_PRE_25', ue_selected_stores),
        (ue_post_25_path, 'UE_POST_25', ue_selected_stores),
    ]
    
    for file_path, file_key, selected_stores in ue_files:
        if not file_path.exists():
            st.warning(f"File not found: {file_path}")
            continue
        try:
            sales, payouts, orders = process_ue_file_for_date_export(file_path, selected_stores)
            # Always add the file entry if at least one metric has data
            if not sales.empty or not payouts.empty or not orders.empty:
                result[file_key] = {
                    'Sales': sales if not sales.empty else pd.DataFrame(),
                    'Payouts': payouts if not payouts.empty else pd.DataFrame(),
                    'Orders': orders if not orders.empty else pd.DataFrame()
                }
            else:
                st.warning(f"No data extracted from {file_path.name} - all pivot tables are empty")
        except Exception as e:
            st.error(f"Error processing {file_path.name}: {str(e)}")
            import traceback
            with st.expander(f"Error details for {file_path.name}"):
                st.code(traceback.format_exc())
    
    return result if result else None


def create_date_export_from_master_files(dd_data_path, ue_data_path, pre_start_date, pre_end_date, post_start_date, post_end_date, excluded_dates=None, operator_name=None):
    """
    Create date-wise exports of DD and UE financial data.
    Creates a single Excel file with 8 sheets (one for each period/platform combination).
    Each sheet contains date-wise data pivoted by Store ID.
    
    Args:
        dd_data_path: Path to DoorDash master file
        ue_data_path: Path to UberEats master file
        pre_start_date: Pre period start date (MM/DD/YYYY)
        pre_end_date: Pre period end date (MM/DD/YYYY)
        post_start_date: Post period start date (MM/DD/YYYY)
        post_end_date: Post period end date (MM/DD/YYYY)
        excluded_dates: List of dates to exclude
    
    Returns:
        Tuple of (excel_bytes, filename) for download
    """
    try:
        # Calculate last year dates
        pre_24_start, pre_24_end = get_last_year_dates(pre_start_date, pre_end_date)
        post_24_start, post_24_end = get_last_year_dates(post_start_date, post_end_date)
        
        # Create a single Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        GAP_COLUMNS = 4
        
        # Process in order: DD_25, UE_25, DD_24, UE_24 (each: Sales, Payouts, Orders)
        if dd_data_path and Path(dd_data_path).exists():
            dd_pre_25 = filter_master_file_by_date_range(Path(dd_data_path), pre_start_date, pre_end_date, DD_DATE_COLUMN_VARIATIONS, excluded_dates)
            dd_post_25 = filter_master_file_by_date_range(Path(dd_data_path), post_start_date, post_end_date, DD_DATE_COLUMN_VARIATIONS, excluded_dates)
            dd_pre_24 = filter_master_file_by_date_range(Path(dd_data_path), pre_24_start, pre_24_end, DD_DATE_COLUMN_VARIATIONS, excluded_dates)
            dd_post_24 = filter_master_file_by_date_range(Path(dd_data_path), post_24_start, post_24_end, DD_DATE_COLUMN_VARIATIONS, excluded_dates)
        else:
            dd_pre_25 = dd_post_25 = dd_pre_24 = dd_post_24 = pd.DataFrame()
        if ue_data_path and Path(ue_data_path).exists():
            ue_pre_25 = filter_master_file_by_date_range(Path(ue_data_path), pre_start_date, pre_end_date, UE_DATE_COLUMN_VARIATIONS, excluded_dates)
            ue_post_25 = filter_master_file_by_date_range(Path(ue_data_path), post_start_date, post_end_date, UE_DATE_COLUMN_VARIATIONS, excluded_dates)
            ue_pre_24 = filter_master_file_by_date_range(Path(ue_data_path), pre_24_start, pre_24_end, UE_DATE_COLUMN_VARIATIONS, excluded_dates)
            ue_post_24 = filter_master_file_by_date_range(Path(ue_data_path), post_24_start, post_24_end, UE_DATE_COLUMN_VARIATIONS, excluded_dates)
        else:
            ue_pre_25 = ue_post_25 = ue_pre_24 = ue_post_24 = pd.DataFrame()
        
        # Sheet order: DD_25, UE_25 (all 25s), then DD_24, UE_24 (all 24s)
        def add_dd_sheets(sheet_label, pre_df, post_df, payout_col_name):
            pre_sales, pre_payouts, pre_orders = _build_period_pivots(pre_df, 'DD', 'Merchant store ID', 'Subtotal', payout_col_name, 'DoorDash order ID')
            post_sales, post_payouts, post_orders = _build_period_pivots(post_df, 'DD', 'Merchant store ID', 'Subtotal', payout_col_name, 'DoorDash order ID')
            _add_pre_post_sheet(wb, f"{sheet_label}_Sales", pre_sales, post_sales, GAP_COLUMNS)
            _add_pre_post_sheet(wb, f"{sheet_label}_Payouts", pre_payouts, post_payouts, GAP_COLUMNS)
            _add_pre_post_sheet(wb, f"{sheet_label}_Orders", pre_orders, post_orders, GAP_COLUMNS)
        def add_ue_sheets(sheet_label, pre_df, post_df):
            ref_df = post_df if pre_df.empty else pre_df
            if ref_df.empty:
                return
            ref_norm, store_col = normalize_store_id_column(ref_df.copy())
            pre_df_norm = normalize_store_id_column(pre_df.copy())[0] if not pre_df.empty else pre_df
            post_df_norm = normalize_store_id_column(post_df.copy())[0] if not post_df.empty else post_df
            pre_sales, pre_payouts, pre_orders = _build_period_pivots(pre_df_norm, 'UE', store_col, 'Sales (excl. tax)', 'Total payout', 'Order ID')
            post_sales, post_payouts, post_orders = _build_period_pivots(post_df_norm, 'UE', store_col, 'Sales (excl. tax)', 'Total payout', 'Order ID')
            _add_pre_post_sheet(wb, f"{sheet_label}_Sales", pre_sales, post_sales, GAP_COLUMNS)
            _add_pre_post_sheet(wb, f"{sheet_label}_Payouts", pre_payouts, post_payouts, GAP_COLUMNS)
            _add_pre_post_sheet(wb, f"{sheet_label}_Orders", pre_orders, post_orders, GAP_COLUMNS)
        
        add_dd_sheets('DD_25', dd_pre_25, dd_post_25, 'Net total')
        add_ue_sheets('UE_25', ue_pre_25, ue_post_25)
        add_dd_sheets('DD_24', dd_pre_24, dd_post_24, 'Net total (for historical reference only)')
        add_ue_sheets('UE_24', ue_pre_24, ue_post_24)
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        tag = (operator_name.strip() if operator_name and isinstance(operator_name, str) and operator_name.strip() else None)
        filename = f"{tag}_date_export_{timestamp}.xlsx" if tag else f"date_export_{timestamp}.xlsx"
        
        return excel_buffer.read(), filename
    
    except Exception as e:
        st.error(f"Error creating date export: {str(e)}")
        import traceback
        with st.expander("Error details"):
            st.code(traceback.format_exc())
        return None, None


def _build_period_pivots(df, platform, store_col, sales_col, payout_col, order_col):
    """
    Build Sales, Payouts, and Orders pivot DataFrames for a single period (Date + store columns).
    Returns (sales_pivot_df, payouts_pivot_df, orders_pivot_df); each may be empty.
    """
    from utils import find_date_column, DD_DATE_COLUMN_VARIATIONS
    empty = pd.DataFrame()
    if df is None or df.empty or store_col is None or store_col not in df.columns:
        return empty.copy(), empty.copy(), empty.copy()
    
    date_col = None
    if platform == 'DD':
        date_col = find_date_column(df, DD_DATE_COLUMN_VARIATIONS)
    else:
        if len(df.columns) > 8:
            date_col = df.columns[8]
    if date_col is None:
        return empty.copy(), empty.copy(), empty.copy()
    
    df = df.copy()
    original_dates = df[date_col].copy()
    if platform == 'UE':
        df[date_col] = pd.to_datetime(df[date_col], format='%m/%d/%Y', errors='coerce')
        if df[date_col].isna().any():
            mask_na = df[date_col].isna()
            df.loc[mask_na, date_col] = pd.to_datetime(original_dates.loc[mask_na], errors='coerce')
    else:
        df[date_col] = pd.to_datetime(df[date_col], format='%m/%d/%Y', errors='coerce')
        if df[date_col].isna().all():
            df[date_col] = pd.to_datetime(original_dates, format='%Y-%m-%d', errors='coerce')
        if df[date_col].isna().all():
            df[date_col] = pd.to_datetime(original_dates, errors='coerce')
    df = df.dropna(subset=[date_col, store_col])
    if df.empty:
        return empty.copy(), empty.copy(), empty.copy()
    
    if sales_col in df.columns:
        df[sales_col] = pd.to_numeric(df[sales_col], errors='coerce').fillna(0)
    # DD_24 must use "Net total (for historical reference only)"; DD_25 uses "Net total".
    # Only fallback for DD_25 when "Net total" is missing (not for DD_24).
    effective_payout_col = payout_col
    if platform == 'DD' and payout_col not in df.columns and payout_col == 'Net total':
        if 'Net total (for historical reference only)' in df.columns:
            effective_payout_col = 'Net total (for historical reference only)'
    if effective_payout_col in df.columns:
        df[effective_payout_col] = pd.to_numeric(df[effective_payout_col], errors='coerce').fillna(0)
    
    sales_agg = df.groupby([date_col, store_col])[sales_col].sum().reset_index() if sales_col in df.columns else pd.DataFrame()
    payouts_agg = df.groupby([date_col, store_col])[effective_payout_col].sum().reset_index() if effective_payout_col in df.columns else pd.DataFrame()
    orders_agg = df.groupby([date_col, store_col])[order_col].nunique().reset_index() if order_col in df.columns else pd.DataFrame()
    
    def _pivot(agg, date_col, store_col, value_col):
        if agg.empty or value_col not in agg.columns:
            return empty.copy()
        p = agg.pivot_table(index=date_col, columns=store_col, values=value_col, aggfunc='sum', fill_value=0)
        p.index = p.index.strftime('%Y-%m-%d')
        p.index.name = 'Date'
        return p.reset_index()
    
    sales_pivot = _pivot(sales_agg, date_col, store_col, sales_col)
    payouts_pivot = _pivot(payouts_agg, date_col, store_col, effective_payout_col)
    orders_pivot = _pivot(orders_agg, date_col, store_col, order_col)
    return sales_pivot, payouts_pivot, orders_pivot


def _add_totals_to_pivot(pivot_df):
    """
    Add a Total column (sum of each row) and a Total row (sum of each column).
    First column is treated as label (Date); rest are numeric.
    Returns a new DataFrame.
    """
    if pivot_df is None or pivot_df.empty:
        return pivot_df
    df = pivot_df.copy()
    label_col = df.columns[0]
    numeric_cols = [c for c in df.columns if c != label_col]
    if not numeric_cols:
        return df
    df['Total'] = df[numeric_cols].sum(axis=1)
    total_row = {label_col: 'Total'}
    for c in numeric_cols:
        total_row[c] = df[c].sum()
    total_row['Total'] = df['Total'].sum()
    order = [label_col] + numeric_cols + ['Total']
    df = df[order]
    total_row_ordered = {c: total_row[c] for c in order}
    df = pd.concat([df, pd.DataFrame([total_row_ordered])], ignore_index=True)
    return df


def _add_pre_post_sheet(wb, sheet_name, pre_pivot, post_pivot, gap_cols=4):
    """
    Create one sheet with Pre data (left), gap_cols empty columns, then Post data (right).
    Adds Total column and Total row to each block.
    """
    from openpyxl.utils.dataframe import dataframe_to_rows
    pre_pivot = _add_totals_to_pivot(pre_pivot) if pre_pivot is not None and not pre_pivot.empty else pre_pivot
    post_pivot = _add_totals_to_pivot(post_pivot) if post_pivot is not None and not post_pivot.empty else post_pivot
    ws = wb.create_sheet(sheet_name)
    start_col_post = 1
    if pre_pivot is not None and not pre_pivot.empty:
        pre_cols = pre_pivot.shape[1]
        pre_rows = 1 + len(pre_pivot)
        for row_idx, row in enumerate(dataframe_to_rows(pre_pivot, index=False, header=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1 or row_idx == pre_rows:
                    cell.font = Font(bold=True)
                if row_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        start_col_post = pre_cols + 1 + gap_cols
    if post_pivot is not None and not post_pivot.empty:
        post_rows = 1 + len(post_pivot)
        for row_idx, row in enumerate(dataframe_to_rows(post_pivot, index=False, header=True), start=1):
            for col_idx, value in enumerate(row, start=start_col_post):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1 or row_idx == post_rows:
                    cell.font = Font(bold=True)
                if row_idx == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')


def _add_period_sheets_to_workbook(wb, df, platform, period_name, store_col, sales_col, payout_col, order_col):
    """
    Add Sales, Payouts, and Orders sheets for a specific period to an existing workbook.
    
    Args:
        wb: openpyxl Workbook object
        df: DataFrame with data
        platform: 'DD' or 'UE'
        period_name: Period name like 'DD_Pre_25', 'UE_Post_24', etc.
        store_col: Name of store ID column
        sales_col: Name of sales column
        payout_col: Name of payout column
        order_col: Name of order ID column
    """
    try:
        # Find date column
        from utils import find_date_column, DD_DATE_COLUMN_VARIATIONS
        date_col = None
        if platform == 'DD':
            date_col = find_date_column(df, DD_DATE_COLUMN_VARIATIONS)
        else:  # UE - hardcode to 9th column (index 8)
            if len(df.columns) > 8:
                date_col = df.columns[8]
            else:
                return
        
        if date_col is None or store_col is None or store_col not in df.columns:
            return
        
        # Convert date column - Store original values before parsing
        original_dates = df[date_col].copy()
        if platform == 'UE':
            # UE files always use MM/DD/YYYY format
            df[date_col] = pd.to_datetime(df[date_col], format='%m/%d/%Y', errors='coerce')
            # Fall back to auto parsing only if format parsing fails
            if df[date_col].isna().any():
                mask_na = df[date_col].isna()
                df.loc[mask_na, date_col] = pd.to_datetime(original_dates.loc[mask_na], errors='coerce')
        else:
            # DD files: Try MM/DD/YYYY format first (most common), then YYYY-MM-DD
            df[date_col] = pd.to_datetime(df[date_col], format='%m/%d/%Y', errors='coerce')
            if df[date_col].isna().all():
                # If all failed, try YYYY-MM-DD format using original values
                df[date_col] = pd.to_datetime(original_dates, format='%Y-%m-%d', errors='coerce')
            # Fall back to auto parsing if format doesn't match
            if df[date_col].isna().all():
                df[date_col] = pd.to_datetime(original_dates, errors='coerce')
        df = df.dropna(subset=[date_col, store_col])
        
        if df.empty:
            return
        
        # Convert to numeric
        if sales_col in df.columns:
            df[sales_col] = pd.to_numeric(df[sales_col], errors='coerce').fillna(0)
        if payout_col in df.columns:
            df[payout_col] = pd.to_numeric(df[payout_col], errors='coerce').fillna(0)
        
        # Aggregate by date and store
        sales_agg = df.groupby([date_col, store_col])[sales_col].sum().reset_index() if sales_col in df.columns else pd.DataFrame()
        payouts_agg = df.groupby([date_col, store_col])[payout_col].sum().reset_index() if payout_col in df.columns else pd.DataFrame()
        orders_agg = df.groupby([date_col, store_col])[order_col].nunique().reset_index() if order_col in df.columns else pd.DataFrame()
        
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Sheet 1: Sales
        if not sales_agg.empty:
            sales_pivot = sales_agg.pivot_table(index=date_col, columns=store_col, values=sales_col, aggfunc='sum', fill_value=0)
            sales_pivot.index = sales_pivot.index.strftime('%Y-%m-%d')
            sales_pivot.index.name = 'Date'
            sales_pivot = sales_pivot.reset_index()
            
            ws_sales = wb.create_sheet(f"{period_name}_Sales")
            for r in dataframe_to_rows(sales_pivot, index=False, header=True):
                ws_sales.append(r)
            # Format header row
            for cell in ws_sales[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sheet 2: Payouts
        if not payouts_agg.empty:
            payouts_pivot = payouts_agg.pivot_table(index=date_col, columns=store_col, values=payout_col, aggfunc='sum', fill_value=0)
            payouts_pivot.index = payouts_pivot.index.strftime('%Y-%m-%d')
            payouts_pivot.index.name = 'Date'
            payouts_pivot = payouts_pivot.reset_index()
            
            ws_payouts = wb.create_sheet(f"{period_name}_Payouts")
            for r in dataframe_to_rows(payouts_pivot, index=False, header=True):
                ws_payouts.append(r)
            # Format header row
            for cell in ws_payouts[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sheet 3: Orders
        if not orders_agg.empty:
            orders_pivot = orders_agg.pivot_table(index=date_col, columns=store_col, values=order_col, aggfunc='sum', fill_value=0)
            orders_pivot.index = orders_pivot.index.strftime('%Y-%m-%d')
            orders_pivot.index.name = 'Date'
            orders_pivot = orders_pivot.reset_index()
            
            ws_orders = wb.create_sheet(f"{period_name}_Orders")
            for r in dataframe_to_rows(orders_pivot, index=False, header=True):
                ws_orders.append(r)
            # Format header row
            for cell in ws_orders[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    except Exception as e:
        st.warning(f"Error adding sheets for {period_name}: {str(e)}")
        import traceback
        st.error(traceback.format_exc())


def _create_period_excel_file(df, platform, period_name, store_col, sales_col, payout_col, order_col):
    """
    Create a single Excel file for a specific period with 3 sheets: Sales, Payouts, Orders.
    
    Args:
        df: DataFrame with data
        platform: 'DD' or 'UE'
        period_name: Period name like 'DD_Pre_25', 'UE_Post_24', etc.
        store_col: Name of store ID column
        sales_col: Name of sales column
        payout_col: Name of payout column
        order_col: Name of order ID column
    
    Returns:
        Bytes of Excel file
    """
    try:
        # Find date column
        from utils import find_date_column, DD_DATE_COLUMN_VARIATIONS
        date_col = None
        if platform == 'DD':
            date_col = find_date_column(df, DD_DATE_COLUMN_VARIATIONS)
        else:  # UE - hardcode to 9th column (index 8)
            if len(df.columns) > 8:
                date_col = df.columns[8]
            else:
                return None
        
        if date_col is None or store_col is None or store_col not in df.columns:
            return None
        
        # Convert date column - Store original values before parsing
        original_dates = df[date_col].copy()
        if platform == 'UE':
            # UE files always use MM/DD/YYYY format
            df[date_col] = pd.to_datetime(df[date_col], format='%m/%d/%Y', errors='coerce')
            # Fall back to auto parsing only if format parsing fails
            if df[date_col].isna().any():
                mask_na = df[date_col].isna()
                df.loc[mask_na, date_col] = pd.to_datetime(original_dates.loc[mask_na], errors='coerce')
        else:
            # DD files: Try MM/DD/YYYY format first (most common), then YYYY-MM-DD
            df[date_col] = pd.to_datetime(df[date_col], format='%m/%d/%Y', errors='coerce')
            if df[date_col].isna().all():
                # If all failed, try YYYY-MM-DD format using original values
                df[date_col] = pd.to_datetime(original_dates, format='%Y-%m-%d', errors='coerce')
            # Fall back to auto parsing if format doesn't match
            if df[date_col].isna().all():
                df[date_col] = pd.to_datetime(original_dates, errors='coerce')
        df = df.dropna(subset=[date_col, store_col])
        
        if df.empty:
            return None
        
        # Convert to numeric
        if sales_col in df.columns:
            df[sales_col] = pd.to_numeric(df[sales_col], errors='coerce').fillna(0)
        if payout_col in df.columns:
            df[payout_col] = pd.to_numeric(df[payout_col], errors='coerce').fillna(0)
        
        # Aggregate by date and store
        sales_agg = df.groupby([date_col, store_col])[sales_col].sum().reset_index() if sales_col in df.columns else pd.DataFrame()
        payouts_agg = df.groupby([date_col, store_col])[payout_col].sum().reset_index() if payout_col in df.columns else pd.DataFrame()
        orders_agg = df.groupby([date_col, store_col])[order_col].nunique().reset_index() if order_col in df.columns else pd.DataFrame()
        
        # Create Excel workbook with 3 sheets
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Sheet 1: Sales
        if not sales_agg.empty:
            sales_pivot = sales_agg.pivot_table(index=date_col, columns=store_col, values=sales_col, aggfunc='sum', fill_value=0)
            sales_pivot.index = sales_pivot.index.strftime('%Y-%m-%d')
            sales_pivot.index.name = 'Date'
            sales_pivot = sales_pivot.reset_index()
            
            ws_sales = wb.create_sheet("Sales")
            for r in dataframe_to_rows(sales_pivot, index=False, header=True):
                ws_sales.append(r)
            # Format header row
            for cell in ws_sales[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sheet 2: Payouts
        if not payouts_agg.empty:
            payouts_pivot = payouts_agg.pivot_table(index=date_col, columns=store_col, values=payout_col, aggfunc='sum', fill_value=0)
            payouts_pivot.index = payouts_pivot.index.strftime('%Y-%m-%d')
            payouts_pivot.index.name = 'Date'
            payouts_pivot = payouts_pivot.reset_index()
            
            ws_payouts = wb.create_sheet("Payouts")
            for r in dataframe_to_rows(payouts_pivot, index=False, header=True):
                ws_payouts.append(r)
            # Format header row
            for cell in ws_payouts[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Sheet 3: Orders
        if not orders_agg.empty:
            orders_pivot = orders_agg.pivot_table(index=date_col, columns=store_col, values=order_col, aggfunc='sum', fill_value=0)
            orders_pivot.index = orders_pivot.index.strftime('%Y-%m-%d')
            orders_pivot.index.name = 'Date'
            orders_pivot = orders_pivot.reset_index()
            
            ws_orders = wb.create_sheet("Orders")
            for r in dataframe_to_rows(orders_pivot, index=False, header=True):
                ws_orders.append(r)
            # Format header row
            for cell in ws_orders[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.read()
    
    except Exception as e:
        st.warning(f"Error creating Excel file for {period_name}: {str(e)}")
        return None
